import datetime
from pathlib import Path
from openpyxl import load_workbook
from . import utils

def run():
    print("\n=== Bulk Vendor Import (Excel) ===\n")
    excel_path = input("Path to Excel file (e.g. inputs/vendors_seed.xlsx): ").strip()

    wb = load_workbook(excel_path)
    ws = wb.active  # assume first sheet is the data

    # expected columns in row 1:
    # org_id | vendor_name | service_description | regulatory_scope |
    # business_owner | likelihood | impact |
    # ac_iam | encrypt | logging | bcp | privacy

    headers = [cell.value for cell in ws[1]]
    col_index = {h: i for i, h in enumerate(headers)}

    required_cols = [
        "org_id", "vendor_name", "service_description",
        "regulatory_scope", "business_owner",
        "likelihood", "impact",
        "ac_iam", "encrypt", "logging", "bcp", "privacy"
    ]

    for rc in required_cols:
        if rc not in col_index:
            print(f"❌ Missing column in Excel: {rc}")
            return

    db = utils.load_vendor_db()

    for row in ws.iter_rows(min_row=2, values_only=True):
        org_id = str(row[col_index["org_id"]]).strip()
        vendor_name = str(row[col_index["vendor_name"]]).strip()
        service_desc = str(row[col_index["service_description"]]).strip()
        regulator = str(row[col_index["regulatory_scope"]]).strip()
        owner = str(row[col_index["business_owner"]]).strip()
        likelihood = str(row[col_index["likelihood"]]).strip().lower()
        impact = str(row[col_index["impact"]]).strip().lower()

        control_scores = {
            "Access Control / Identity Management": int(row[col_index["ac_iam"]]),
            "Encryption & Key Management": int(row[col_index["encrypt"]]),
            "Monitoring & Logging": int(row[col_index["logging"]]),
            "Business Continuity / DR / Resilience": int(row[col_index["bcp"]]),
            "Privacy & Regulatory Compliance": int(row[col_index["privacy"]]),
        }

        overall_control = round(
            sum(control_scores.values()) / len(control_scores), 2
        )

        risk_bucket = utils.classify_risk_bucket(likelihood, impact)

        record = {
            "org_id": org_id,
            "vendor_name": vendor_name,
            "service": service_desc,
            "regulator": regulator,
            "business_owner": owner,
            "likelihood": likelihood,
            "impact": impact,
            "risk_bucket": risk_bucket,
            "control_scores": control_scores,
            "overall_control_score": overall_control,
            "assessed_at": datetime.datetime.now().isoformat(),
            "assessed_by": "import",
            "open_actions": [],
            "risk_acceptances": []
        }

        missing = utils.validate_vendor_record(record)
        if missing:
            print(f"⚠ Skipping {vendor_name} for {org_id} due to missing {missing}")
            continue

        # commit to DB
        db = utils.ensure_org(db, org_id)
        db[org_id][vendor_name] = record
        utils.snapshot_history(org_id, vendor_name, record)

    utils.save_vendor_db(db)
    print("✅ Import complete. Vendors added/updated in data/vendors.json.\n")
