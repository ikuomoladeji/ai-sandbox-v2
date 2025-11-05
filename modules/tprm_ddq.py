

import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from . import utils
from modules.api_client import ask_model

# === WEIGHT MODEL (TOTAL = 100%)
TPRM_DOMAINS = [
    {
        "name": "Data Protection",
        "weight": 25,
        "questions": [
            {"q": "Is customer data encrypted at rest and in transit?", "weight": 1.0},
            {"q": "Are there defined data retention and disposal procedures?", "weight": 1.0},
            {"q": "Is there evidence of SOC2/ISO27001 certifications related to data protection?", "weight": 1.0},
        ],
    },
    {
        "name": "Compliance",
        "weight": 15,
        "questions": [
            {"q": "Do you maintain compliance with GDPR, HIPAA, PCI-DSS or similar?", "weight": 1.0},
            {"q": "Are certifications / audits current and valid (not expired)?", "weight": 1.0},
        ],
    },
    {
        "name": "Access Control",
        "weight": 15,
        "questions": [
            {"q": "Is MFA enforced for privileged and remote access?", "weight": 1.0},
            {"q": "Are access reviews / offboarding revocations performed regularly?", "weight": 1.0},
        ],
    },
    {
        "name": "Incident Response",
        "weight": 15,
        "questions": [
            {"q": "Is there a documented incident response plan?", "weight": 1.0},
            {"q": "Has the IR plan been tested in the last 12 months?", "weight": 1.0},
            {"q": "Is there a defined escalation / RACI for incidents?", "weight": 1.0},
        ],
    },
    {
        "name": "Business Continuity",
        "weight": 10,
        "questions": [
            {"q": "Do you have BCP/DR plans that are tested annually?", "weight": 1.0},
            {"q": "Are RTO/RPO objectives defined and achievable for this service?", "weight": 1.0},
        ],
    },
    {
        "name": "Subprocessor Management",
        "weight": 10,
        "questions": [
            {"q": "Do you assess and monitor critical subprocessors / 4th parties?", "weight": 1.0},
            {"q": "Do you maintain exit / contingency plans for critical suppliers?", "weight": 1.0},
        ],
    },
    {
        "name": "Governance & Documentation",
        "weight": 10,
        "questions": [
            {"q": "Are key policies owned, reviewed, and approved annually?", "weight": 1.0},
            {"q": "Are risk/security roles and responsibilities clearly documented?", "weight": 1.0},
        ],
    },
]

def ask_score(question_text):
    print(f"\n{question_text}")
    print("  Rate 1–5:")
    print("   5 = Strong control (evidence verified)")
    print("   4 = Acceptable control (minor gaps)")
    print("   3 = Moderate control (some gaps)")
    print("   2 = Weak control (no formal process)")
    print("   1 = Unacceptable (missing control)")
    while True:
        raw = input("  Score: ").strip()
        try:
            val = int(raw)
            if 1 <= val <= 5:
                return val
        except ValueError:
            pass
        print("  Enter 1, 2, 3, 4, or 5.")

def score_domain(domain):
    weighted_sum = 0.0
    total_w = 0.0
    q_details = []

    print(f"\n--- {domain['name']} (Weight {domain['weight']}%) ---")
    for q in domain["questions"]:
        s = ask_score(q["q"])
        q_details.append({
            "question": q["q"],
            "score": s,
            "sub_weight": q["weight"],
        })
        weighted_sum += s * q["weight"]
        total_w += q["weight"]

    domain_avg = round(weighted_sum / total_w, 2)
    return domain_avg, q_details

def compute_total_weighted_score(domain_results):
    weighted_sum_1to5 = 0.0
    normalized_pct_sum = 0.0
    for d in domain_results:
        w_frac = d["weight_pct"] / 100.0
        weighted_sum_1to5 += d["score"] * w_frac
        normalized_pct_sum += (d["score"] / 5.0) * d["weight_pct"]
    total_weighted_score = round(weighted_sum_1to5, 2)
    composite_pct_score = round(normalized_pct_sum, 2)
    return total_weighted_score, composite_pct_score

def classify_risk_level(total_weighted_score):
    if total_weighted_score >= 4.0:
        return "Low"
    elif total_weighted_score >= 3.0:
        return "Medium"
    else:
        return "High"

def export_to_excel(org, vendor, assessment_date, likelihood, impact,
                    total_weighted_score, risk_level, domain_results,
                    approvals, evidence_notes):

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Risk Assessment"

    bold = Font(bold=True)
    header_bold = Font(bold=True, size=14)
    mid_bold = Font(bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1
    ws["A1"] = f"Vendor Risk Assessment – {vendor} ({org})"
    ws["A1"].font = header_bold
    row += 2

    ws["A"+str(row)] = "Assessment Summary"
    ws["A"+str(row)].font = mid_bold
    row += 1

    summary_fields = [
        ("Organisation", org),
        ("Vendor", vendor),
        ("Assessment Date", assessment_date),
        ("Likelihood", likelihood),
        ("Impact", impact),
        ("Total Weighted Score (1–5 style)", total_weighted_score),
        ("Overall Risk Level", risk_level),
    ]
    for label, value in summary_fields:
        ws["A"+str(row)] = label
        ws["A"+str(row)].font = bold
        ws["B"+str(row)] = value
        row += 1

    row += 1

    ws["A"+str(row)] = "Control Domain Breakdown"
    ws["A"+str(row)].font = mid_bold
    row += 1

    ws["A"+str(row)] = "Control Domain"
    ws["B"+str(row)] = "Weight %"
    ws["C"+str(row)] = "Score (1–5)"
    ws["D"+str(row)] = "Weighted Contribution"
    ws["E"+str(row)] = "Evidence / Notes"
    for col in "ABCDE":
        cell = ws[col+str(row)]
        cell.font = bold
        cell.alignment = center
        cell.border = border
    row += 1

    for d in domain_results:
        contrib = round(d["score"] * (d["weight_pct"]/100.0), 2)
        ws["A"+str(row)] = d["name"]
        ws["B"+str(row)] = d["weight_pct"]
        ws["C"+str(row)] = d["score"]
        ws["D"+str(row)] = contrib
        ws["E"+str(row)] = evidence_notes.get(d["name"], "")
        for col in "ABCDE":
            ws[col+str(row)].alignment = wrap
            ws[col+str(row)].border = border
        row += 1

    row += 2

    ws["A"+str(row)] = "Final Approval / Sign-off"
    ws["A"+str(row)].font = mid_bold
    row += 1

    ws["A"+str(row)] = "Reviewer"
    ws["B"+str(row)] = "Role"
    ws["C"+str(row)] = "Decision"
    ws["D"+str(row)] = "Signature / Notes"
    for col in "ABCD":
        cell = ws[col+str(row)]
        cell.font = bold
        cell.alignment = center
        cell.border = border
    row += 1

    for ap in approvals:
        ws["A"+str(row)] = ap.get("reviewer","")
        ws["B"+str(row)] = ap.get("role","")
        ws["C"+str(row)] = ap.get("decision","")
        ws["D"+str(row)] = ap.get("notes","")
        for col in "ABCD":
            ws[col+str(row)].alignment = wrap
            ws[col+str(row)].border = border
        row += 1

    widths = [("A",32), ("B",16), ("C",14), ("D",22), ("E",44)]
    for col_letter, w in widths:
        ws.column_dimensions[col_letter].width = w

    out_dir = Path("outputs")
    out_dir.mkdir(exist_ok=True)
    safe_org = org.replace(" ", "_")
    safe_vendor = vendor.replace(" ", "_")
    filename = f"{safe_org}_{safe_vendor}_assessment.xlsx"
    out_path = out_dir / filename
    wb.save(out_path)

    print(f"\n📄 Excel exported: {out_path.resolve()}")


# ---------- helper menu utilities ----------

def _list_orgs(db):
    return list(db.keys())

def _list_vendors_for_org(db, org):
    if org in db:
        return list(db[org].keys())
    return []

def _pretty_print_org_vendors(db, org):
    print(f"\nVendors for organisation: {org}")
    print("--------------------------------------------------")
    if org not in db or not db[org]:
        print("No vendors recorded yet.")
        return
    for v_name, rec in db[org].items():
        w = rec.get("weighted_score", "n/a")
        rl = rec.get("risk_level", "n/a")
        date = rec.get("assessment_date", rec.get("assessed_at", "n/a"))
        print(f"- {v_name} | score={w} | risk={rl} | last={date}")
    print("--------------------------------------------------\n")

def _edit_existing_record(db, org, vendor):
    record = db[org][vendor]
    print(f"\nEditing existing vendor '{vendor}' in org '{org}'")
    print("Current profile:")
    print(f"  Service: {record.get('service','')}")
    print(f"  Business owner: {record.get('business_owner','')}")
    print(f"  Regulator: {record.get('regulator','')}")
    print(f"  Likelihood: {record.get('likelihood','')}")
    print(f"  Impact: {record.get('impact','')}")
    print(f"  Weighted score: {record.get('weighted_score','')}")
    print(f"  Risk level: {record.get('risk_level','')}")
    print("--------------------------------------------------")

    new_service = input("Update service / data handled (leave blank to keep): ").strip()
    new_owner = input("Update business owner / sponsor (leave blank to keep): ").strip()
    new_reg = input("Update regulatory scope (leave blank to keep): ").strip()
    new_like = input("Update likelihood (low/medium/high, blank=keep): ").strip()
    new_imp  = input("Update impact (low/medium/high, blank=keep): ").strip()

    if new_service:
        record["service"] = new_service
    if new_owner:
        record["business_owner"] = new_owner
    if new_reg:
        record["regulator"] = new_reg
    if new_like:
        record["likelihood"] = new_like
    if new_imp:
        record["impact"] = new_imp

    rerun = input("\nRe-run DDQ scoring and regenerate outputs? (y/n): ").strip().lower()
    if rerun == "y":
        return "rescore"

    db[org][vendor] = record
    utils.save_vendor_db(db)
    utils.snapshot_history(org, vendor, record)

    print("\n✅ Vendor metadata updated and snapshot stored (no DDQ re-score).")
    return "done"


# ---------- main workflow ----------

def run(model_name):
    db = utils.load_vendor_db()

    print("\n=== Vendor Assessment Menu ===")
    print("1) Start NEW assessment for a vendor")
    print("2) Update / continue EXISTING vendor")
    print("3) List vendors for an organisation (view only)")
    print("4) Cancel / go back")
    mode_choice = input("Choose 1 / 2 / 3 / 4: ").strip()

    # VIEW ONLY
    if mode_choice == "3":
        if not db:
            print("\nNo organisations found yet.")
            return
        print("\nKnown organisations:")
        org_names = _list_orgs(db)
        for idx, org_name in enumerate(org_names, start=1):
            print(f"{idx}. {org_name}")
        pick = input("Which organisation number to view? ").strip()
        try:
            pick_idx = int(pick) - 1
        except:
            print("Invalid number.")
            return
        if pick_idx < 0 or pick_idx >= len(org_names):
            print("Out of range.")
            return
        org = org_names[pick_idx]
        _pretty_print_org_vendors(db, org)
        return

    # EXISTING VENDOR
    if mode_choice == "2":
        if not db:
            print("\nNo organisations found yet.")
            return

        print("\nKnown organisations:")
        org_names = _list_orgs(db)
        for idx, org_name in enumerate(org_names, start=1):
            print(f"{idx}. {org_name}")
        pick_org = input("Select organisation number: ").strip()
        try:
            org_idx = int(pick_org) - 1
        except:
            print("Invalid number.")
            return
        if org_idx < 0 or org_idx >= len(org_names):
            print("Out of range.")
            return
        org = org_names[org_idx]

        vendors = _list_vendors_for_org(db, org)
        if not vendors:
            print("\nNo vendors recorded for that organisation yet.")
            return
        print(f"\nVendors under {org}:")
        for vidx, vname in enumerate(vendors, start=1):
            print(f"{vidx}. {vname}")
        pick_vendor = input("Select vendor number: ").strip()
        try:
            vend_idx = int(pick_vendor) - 1
        except:
            print("Invalid number.")
            return
        if vend_idx < 0 or vend_idx >= len(vendors):
            print("Out of range.")
            return
        vendor = vendors[vend_idx]

        decision = _edit_existing_record(db, org, vendor)
        if decision != "rescore":
            # user only updated metadata + saved
            return

        # RESCORE path - preload
        existing_record = db[org][vendor]
        service_desc = existing_record.get("service","")
        owner = existing_record.get("business_owner","")
        regulator = existing_record.get("regulator","")
        likelihood = existing_record.get("likelihood","").lower()
        impact = existing_record.get("impact","").lower()
        assessor = input("Your name / initials: ").strip() or existing_record.get("assessed_by","")

    # CANCEL
    elif mode_choice == "4":
        print("Cancelled.")
        return

    # NEW VENDOR
    else:
        org = input("\nOrganisation / Client name: ").strip()
        if org not in db:
            db[org] = {}

        vendor = input("Vendor name: ").strip()
        if vendor not in db[org]:
            db[org][vendor] = {
                "organisation": org,
                "vendor_name": vendor,
                "created": datetime.datetime.now().isoformat(),
            }

        service_desc = input("Describe the service / data handled: ").strip()
        owner = input("Business owner / sponsor: ").strip()
        regulator = input("Regulatory scope (GDPR, PCI, HIPAA, etc.): ").strip()
        likelihood = input("Likelihood (low/medium/high): ").strip().lower()
        impact = input("Impact (low/medium/high): ").strip().lower()
        assessor = input("Who is performing this assessment (your name/initials)?: ").strip()

    # ----- DDQ SCORING -----
    domain_results = []
    for d in TPRM_DOMAINS:
        domain_score, q_details = score_domain(d)
        domain_results.append({
            "name": d["name"],
            "weight_pct": d["weight"],
            "score": domain_score,
            "questions": q_details,
        })

    total_weighted_score, composite_pct_score = compute_total_weighted_score(domain_results)
    risk_level = classify_risk_level(total_weighted_score)

    print("\n===== WEIGHTED RESULT =====")
    for r in domain_results:
        contrib = round(r["score"] * (r["weight_pct"]/100.0), 2)
        print(f"- {r['name']}: {r['score']} × {r['weight_pct']}% → {contrib}")
    print(f"\nTotal Weighted Score (1–5 style): {total_weighted_score}")
    print(f"Overall Risk Level: {risk_level}")
    print(f"Composite % Score (dashboard style): {composite_pct_score}/100\n")

    # ----- EVIDENCE NOTES -----
    evidence_notes = {}
    print("Add evidence / notes per domain (e.g. 'SOC2 Type II valid to Jun 2025').")
    for r in domain_results:
        note = input(f"Evidence for {r['name']}: ").strip()
        evidence_notes[r["name"]] = note

    # ----- APPROVALS -----
    approvals = []
    print("\nApprovals / Sign-off (IT, Risk, Procurement). Leave reviewer blank to finish.")
    while True:
        reviewer = input("Reviewer name (blank to finish): ").strip()
        if reviewer == "":
            break
        role = input("  Role / function: ").strip()
        decision = input("  Decision (Approved / Mitigate / Renew / etc.): ").strip()
        notes = input("  Notes / signature placeholder: ").strip()
        approvals.append({
            "reviewer": reviewer,
            "role": role,
            "decision": decision,
            "notes": notes
        })
        print("Added.\n")

    # ----- BUILD / SAVE RECORD -----
    now_iso = datetime.datetime.now().isoformat()
    today_str = str(datetime.date.today())

    record = db[org][vendor]
    record.update({
        "assessment_date": today_str,
        "assessed_at": now_iso,
        "assessed_by": assessor,
        "service": service_desc,
        "business_owner": owner,
        "regulator": regulator,
        "likelihood": likelihood,
        "impact": impact,
        "domains": domain_results,
        "weighted_score": total_weighted_score,
        "risk_level": risk_level,
        "composite_pct_score": composite_pct_score,
        "evidence_notes": evidence_notes,
        "approvals": approvals,
    })

    db[org][vendor] = record
    utils.save_vendor_db(db)
    utils.snapshot_history(org, vendor, record)

    # ----- CHOOSE OUTPUT BEHAVIOUR -----
    print("\nWhat do you want to do next?")
    print("1) Generate Excel pack + narrative report for THIS vendor now")
    print("2) Just save data (no reports yet)")
    print("3) Save + generate vendor treatment summary (board-style for THIS vendor)")
    print("4) Save + generate ORG-LEVEL management summary (ALL vendors in this org)")
    print("5) Stop here (saved, no output)")
    next_action = input("Choose 1 / 2 / 3 / 4 / 5: ").strip()

    if next_action in ["2", "5"]:
        print("\n💾 Vendor data saved.")
        print("📝 No reports generated right now.")
        print("You can still generate outputs later via main menu (options 2, 4, or 7).")
        print("\n✅ Done.\n")
        return

    # build narrative text + per-vendor Excel first (for 1 / 3 / 4)
    narrative_lines = []
    narrative_lines.append(f"Vendor: {vendor}")
    narrative_lines.append(f"Organisation: {org}")
    narrative_lines.append(f"Assessment date: {today_str}")
    narrative_lines.append(f"Weighted Score: {total_weighted_score}")
    narrative_lines.append(f"Risk Level: {risk_level}")
    narrative_lines.append(f"Likelihood: {likelihood} | Impact: {impact}")
    narrative_lines.append("\nDomain breakdown:")
    for r in domain_results:
        narrative_lines.append(
            f"- {r['name']}: score={r['score']} (weight={r['weight_pct']}%)"
        )
    narrative_lines.append("\nEvidence notes:")
    for dom_name, note in evidence_notes.items():
        narrative_lines.append(f"* {dom_name}: {note}")
    narrative_lines.append("\nApprovals:")
    for a in approvals:
        narrative_lines.append(
            f"- {a['reviewer']} ({a['role']}): {a['decision']} | {a['notes']}"
        )

    out_dir = Path("outputs")
    out_dir.mkdir(exist_ok=True)
    safe_org = org.replace(" ", "_")
    safe_vendor = vendor.replace(" ", "_")
    txt_path = out_dir / f"{safe_org}_{safe_vendor}_Risk_Assessment.txt"
    txt_path.write_text("\n".join(narrative_lines), encoding="utf-8")

    export_to_excel(
        org=org,
        vendor=vendor,
        assessment_date=today_str,
        likelihood=likelihood,
        impact=impact,
        total_weighted_score=total_weighted_score,
        risk_level=risk_level,
        domain_results=domain_results,
        approvals=approvals,
        evidence_notes=evidence_notes,
    )

    print(f"\n📝 Narrative report saved: {txt_path.resolve()}")
    print("📊 Excel pack saved in ./outputs/")
    print("💾 Record saved to data/vendors.json and snapshotted in history/")

    # vendor-level treatment snapshot
    if next_action == "3":
        from . import risk_treatment
        print("\n📊 Generating board / treatment summary for this vendor ...\n")
        risk_treatment.run_for_single_vendor(org, vendor, model_name)

    # org-level summary across all vendors
    if next_action == "4":
        from . import risk_treatment
        print(f"\n📊 Generating ORG-LEVEL summary for {org} ...\n")
        risk_treatment.run_for_org(org, model_name)

    print("\n✅ Done. Returning to main menu.\n")
    return
